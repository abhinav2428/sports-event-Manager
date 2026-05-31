"""
Excel import / export endpoints for Swimmers and Teams.

Swimmers
  GET  /swimmers/excel/template  → blank .xlsx with standard headers
  POST /swimmers/excel/import    → upload .xlsx, upsert by roll_number
  GET  /swimmers/excel/export    → download all swimmers as .xlsx

Teams
  GET  /teams/excel/template     → blank .xlsx with standard headers
  POST /teams/excel/import       → upload .xlsx, upsert by name
  GET  /teams/excel/export       → download all teams as .xlsx
"""
from __future__ import annotations

import io
import uuid
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.api.deps import require_admin
from app.models.swimmer import Swimmer, Team
from app.models.user import User

excel_router = APIRouter(tags=["Excel"])

# ── Column definitions ─────────────────────────────────────────

SWIMMER_COLS = [
    "Name",
    "Roll Number",
    "Gender",        # M or F
    "College",
    "Date of Birth", # YYYY-MM-DD  (optional)
    "Email",
    "Phone",
]

TEAM_COLS = [
    "Name",
    "College",
    "Gender",        # M or F
]


# ── Helpers ────────────────────────────────────────────────────

def _make_workbook(headers: list[str]) -> openpyxl.Workbook:
    """Return a workbook with a single styled header row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    # Bold + light-blue header
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="BDD7EE")
    font = Font(bold=True)
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    # Auto column widths
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)
    return wb


def _stream_wb(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_date(value) -> Optional[date]:
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _cell(row, idx: int):
    """Return stripped string or None for a cell value."""
    v = row[idx] if idx < len(row) else None
    return str(v).strip() if v is not None and str(v).strip() else None


# ══════════════════════════════════════════════════════════════
#  SWIMMERS
# ══════════════════════════════════════════════════════════════

@excel_router.get("/swimmers/excel/template", summary="Download blank swimmer template")
def swimmer_template(_: User = Depends(require_admin)):
    wb = _make_workbook(SWIMMER_COLS)
    return _stream_wb(wb, "swimmers_template.xlsx")


@excel_router.post("/swimmers/excel/import", summary="Import swimmers from Excel (upsert by roll number)")
def import_swimmers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx files are supported")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not parse the uploaded file. Make sure it is a valid .xlsx.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    # Validate header
    header = [str(h).strip() if h else "" for h in rows[0]]
    missing = [c for c in ["Name", "Roll Number", "Gender"] if c not in header]
    if missing:
        raise HTTPException(400, f"Missing required columns: {missing}. Expected headers: {SWIMMER_COLS}")

    col = {name: header.index(name) for name in header if name in SWIMMER_COLS}

    created = updated = skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows[1:], start=2):
        name       = _cell(row, col.get("Name", -1))
        roll       = _cell(row, col.get("Roll Number", -1))
        gender_raw = _cell(row, col.get("Gender", -1))

        # Skip fully empty rows
        if not any([name, roll, gender_raw]):
            skipped += 1
            continue

        # Validate required fields
        if not name:
            errors.append(f"Row {i}: Name is required"); continue
        if not roll:
            errors.append(f"Row {i}: Roll Number is required"); continue
        if gender_raw not in ("M", "F", "Male", "Female", "male", "female"):
            errors.append(f"Row {i}: Gender must be M or F, got '{gender_raw}'"); continue

        gender = "M" if gender_raw.upper().startswith("M") else "F"
        college = _cell(row, col.get("College", -1))
        dob     = _parse_date(row[col["Date of Birth"]] if "Date of Birth" in col else None)
        email   = _cell(row, col.get("Email", -1))
        phone   = _cell(row, col.get("Phone", -1))

        existing = db.query(Swimmer).filter(Swimmer.roll_number == roll).first()
        if existing:
            existing.name    = name
            existing.gender  = gender
            existing.college = college
            existing.dob     = dob
            existing.email   = email
            existing.phone   = phone
            updated += 1
        else:
            db.add(Swimmer(
                id=str(uuid.uuid4()),
                name=name,
                roll_number=roll,
                gender=gender,
                college=college,
                dob=dob,
                email=email,
                phone=phone,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@excel_router.get("/swimmers/excel/export", summary="Export all swimmers to Excel")
def export_swimmers(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    swimmers = db.query(Swimmer).order_by(Swimmer.name).all()
    wb = _make_workbook(SWIMMER_COLS)
    ws = wb.active
    for s in swimmers:
        ws.append([
            s.name,
            s.roll_number,
            s.gender,
            s.college or "",
            str(s.dob) if s.dob else "",
            s.email or "",
            s.phone or "",
        ])
    return _stream_wb(wb, "swimmers_export.xlsx")


# ══════════════════════════════════════════════════════════════
#  TEAMS
# ══════════════════════════════════════════════════════════════

@excel_router.get("/teams/excel/template", summary="Download blank team template")
def team_template(_: User = Depends(require_admin)):
    wb = _make_workbook(TEAM_COLS)
    return _stream_wb(wb, "teams_template.xlsx")


@excel_router.post("/teams/excel/import", summary="Import teams from Excel (upsert by name)")
def import_teams(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx files are supported")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not parse the uploaded file. Make sure it is a valid .xlsx.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    header = [str(h).strip() if h else "" for h in rows[0]]
    missing = [c for c in ["Name", "Gender"] if c not in header]
    if missing:
        raise HTTPException(400, f"Missing required columns: {missing}. Expected headers: {TEAM_COLS}")

    col = {name: header.index(name) for name in header if name in TEAM_COLS}

    created = updated = skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows[1:], start=2):
        name       = _cell(row, col.get("Name", -1))
        gender_raw = _cell(row, col.get("Gender", -1))

        if not any([name, gender_raw]):
            skipped += 1
            continue

        if not name:
            errors.append(f"Row {i}: Name is required"); continue
        if gender_raw not in ("M", "F", "Male", "Female", "male", "female"):
            errors.append(f"Row {i}: Gender must be M or F, got '{gender_raw}'"); continue

        gender  = "M" if gender_raw.upper().startswith("M") else "F"
        college = _cell(row, col.get("College", -1))

        existing = db.query(Team).filter(Team.name == name).first()
        if existing:
            existing.college = college
            existing.gender  = gender
            updated += 1
        else:
            db.add(Team(id=str(uuid.uuid4()), name=name, college=college, gender=gender))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@excel_router.get("/teams/excel/export", summary="Export all teams to Excel")
def export_teams(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    teams = db.query(Team).order_by(Team.name).all()
    wb = _make_workbook(TEAM_COLS)
    ws = wb.active
    for t in teams:
        ws.append([t.name, t.college or "", t.gender])
    return _stream_wb(wb, "teams_export.xlsx")
